"""
RCSB PDB report routes.

Search RCSB, view results, export to Excel, sync APS pub database.
Staff-only access.
"""

import io
import json
import logging
import os
import re
import sys
import tempfile
import threading
import time
import uuid
from datetime import datetime, date, timedelta, timezone
from difflib import SequenceMatcher
from typing import List, Optional, Union
from pathlib import Path

import pandas as pd
import requests
from fastapi import APIRouter, Depends, HTTPException, UploadFile, File
from fastapi.responses import StreamingResponse
from pydantic import BaseModel
from sqlalchemy.orm import Session

from qp2.web_app.backend.security import verify_token
from qp2.web_app.backend.auth import is_staff_member

from qp2.rcsb_tools.rcsb_core import (
    Presets, RCSB, Parser, APS_Parser, APS_Pub_Parser, Report,
    doi_url, to_date, date_str, iso_date_str, concat_authors,
    get_date_range, fuzzy_str_compare,
)
from qp2.db import (
    APSPublication, APSSyncStatus,
    ScheduledTaskRecipient, ScheduledJobState,
)
from qp2.web_app.backend.email_utils import send_mail, send_admin_alert
from qp2.log.logging_config import get_logger

logger = get_logger(__name__)

router = APIRouter(prefix="/rcsb", tags=["rcsb"])

# --- Background search job store ---
# In-memory only — sufficient for single-worker deployment.
_jobs: dict = {}
_jobs_lock = threading.Lock()


def _update_job(job_id: str, **kwargs) -> None:
    with _jobs_lock:
        if job_id in _jobs:
            _jobs[job_id].update(kwargs)


def _cleanup_old_jobs() -> None:
    cutoff = time.time() - 3600
    with _jobs_lock:
        stale = [k for k, v in _jobs.items() if v.get("created_at", 0) < cutoff]
        for k in stale:
            del _jobs[k]


# Injected by main.py after DB init. Jobs read this at run time instead of
# receiving the factory as a kwarg — APScheduler's SQLAlchemyJobStore pickles
# kwargs, and SQLAlchemy session/engine objects reference closures
# (e.g. create_engine.<locals>.connect) that cannot be pickled.
_session_factory = None


def set_session_factory(factory):
    global _session_factory
    _session_factory = factory


def _get_session():
    if _session_factory is None:
        raise RuntimeError("rcsb_routes.session_factory not set")
    return _session_factory()


# --- Staff-only dependency ---

def require_staff(user: str = Depends(verify_token)):
    if not is_staff_member(user):
        raise HTTPException(status_code=403, detail="Staff access required")
    return user


# --- DB session dependency (overridden in main.py) ---

def get_db_session():
    raise NotImplementedError("Must be overridden via app.dependency_overrides")


# --- Pydantic models ---

class SearchRequest(BaseModel):
    report_type: str  # gmca, aps, aps_pub, generic
    start_date: Optional[str] = None
    end_date: Optional[str] = None
    search_text: Optional[str] = None
    pdb_codes: Optional[List[str]] = None


# --- Helpers ---

_HYPERLINK_RE = re.compile(r'^=HYPERLINK\("([^"]+)",\s*"[^"]*"\)$')
_EMAIL_RE = re.compile(r"^[^@\s]+@[^@\s]+\.[^@\s]+$")

_SCHEDULED_TASKS = {
    "gmca_weekly": {
        "label": "GMCA Weekly",
        "default_recipients": [
            "qxu@anl.gov", "mbecker@anl.gov",
            "rfischetti@anl.gov", "janetsmi@umich.edu",
        ],
    },
    "aps_pub_monthly": {
        "label": "APS Pub Monthly",
        "default_recipients": [
            "qxu@anl.gov", "rfischetti@anl.gov",
            "kahrens@anl.gov", "jskwarek@anl.gov",
        ],
    },
}


def _seed_scheduled_tasks(session: Session) -> None:
    """Seed default recipients and job state on first startup (idempotent)."""
    for task_name, cfg in _SCHEDULED_TASKS.items():
        existing = session.query(ScheduledTaskRecipient).filter_by(
            task_name=task_name
        ).first()
        if existing is None:
            for email in cfg["default_recipients"]:
                session.add(ScheduledTaskRecipient(task_name=task_name, email=email))
        if session.query(ScheduledJobState).filter_by(task_name=task_name).first() is None:
            session.add(ScheduledJobState(task_name=task_name, last_run_at=None))
    session.commit()
    logger.info("Scheduled tasks seeded")


def _write_excel(df: pd.DataFrame, target: Union[str, io.BytesIO]) -> None:
    """Write DataFrame to .xlsx. Target is a file path string or BytesIO buffer."""
    from openpyxl.utils import get_column_letter

    df = df.copy()
    for col in df.columns:
        if df[col].dtype == "object":
            df[col] = df[col].apply(
                lambda x: x.isoformat() if isinstance(x, date) else x
            )

    writer = pd.ExcelWriter(target, engine="openpyxl")
    df.to_excel(writer, sheet_name="Sheet1", index=False)
    ws = writer.sheets["Sheet1"]

    for idx, col in enumerate(df.columns):
        series = df[col].fillna("").astype(str)
        max_len = min(max(series.str.len().max(), len(str(col))) + 1, 50)
        ws.column_dimensions[get_column_letter(idx + 1)].width = max_len

    for col_name in ("doi", "pubmed"):
        if col_name in df.columns:
            col_idx = df.columns.get_loc(col_name) + 1
            for row_idx, val in enumerate(df[col_name]):
                if pd.notna(val) and str(val).startswith("http"):
                    cell = ws.cell(row=row_idx + 2, column=col_idx)
                    cell.hyperlink = str(val)
                    cell.style = "Hyperlink"

    writer.close()


def _job_gmca_weekly() -> None:
    """APScheduler job: send GMCA weekly report email."""
    session = _get_session()
    try:
        # Race-safe dedup: lock the state row, skip if another worker fired
        # this job within the last 12 h. SELECT ... FOR UPDATE blocks the
        # second worker until the first commits, so the loser sees a fresh
        # last_run_at and exits cleanly.
        state = (session.query(ScheduledJobState)
                 .filter_by(task_name="gmca_weekly")
                 .with_for_update()
                 .first())
        end_dt = datetime.now(timezone.utc).replace(tzinfo=None)
        if state and state.last_run_at \
                and (end_dt - state.last_run_at) < timedelta(hours=12):
            logger.info("gmca_weekly: ran within last 12 h, skipping duplicate fire")
            session.commit()
            return

        recipients = [
            r.email for r in session.query(ScheduledTaskRecipient)
            .filter_by(task_name="gmca_weekly").all()
        ]
        if not recipients:
            logger.warning("gmca_weekly: no recipients configured, skipping")
            session.commit()
            return

        start_dt = state.last_run_at if (state and state.last_run_at) \
            else end_dt - timedelta(days=30)

        params = SearchRequest(
            report_type="gmca",
            start_date=start_dt.strftime("%Y-%m-%d"),
            end_date=end_dt.strftime("%Y-%m-%d"),
        )
        rows, _ = _run_rcsb_query(params)
        today_str = end_dt.strftime("%Y-%m-%d")

        if not rows:
            send_mail(
                subject=f"[GMCA Weekly] No new structures ({today_str})",
                body=(
                    f"No new GMCA structures were released between "
                    f"{start_dt.strftime('%Y-%m-%d')} and {today_str}."
                ),
                to=recipients,
            )
        else:
            df = pd.DataFrame(rows)
            with tempfile.NamedTemporaryFile(suffix=".xlsx", prefix=f"gmca_weekly_{today_str}_", delete=False) as tmp:
                xlsx_path = tmp.name
            _write_excel(df, xlsx_path)
            try:
                send_mail(
                    subject=f"[GMCA Weekly] {len(rows)} new structures ({today_str})",
                    body=(
                        f"GMCA weekly report: {len(rows)} structures released "
                        f"between {start_dt.strftime('%Y-%m-%d')} and {today_str}.\n"
                        f"See attached spreadsheet."
                    ),
                    to=recipients,
                    attachment_path=xlsx_path,
                )
            finally:
                if os.path.exists(xlsx_path):
                    os.unlink(xlsx_path)

        if state is None:
            state = ScheduledJobState(task_name="gmca_weekly")
            session.add(state)
        state.last_run_at = end_dt
        session.commit()
        logger.info(f"gmca_weekly completed: {len(rows)} results")

    except Exception as e:
        session.rollback()
        logger.exception(f"gmca_weekly job failed: {e}")
        send_admin_alert(
            subject=f"[GMCA Weekly] Job failed — {type(e).__name__}",
            body=f"Error during gmca_weekly scheduled job:\n\n{e}",
        )
    finally:
        session.close()


def _job_aps_pub_monthly() -> None:
    """APScheduler job: send APS publication monthly report email."""
    session = _get_session()
    try:
        # Same race-safe dedup as gmca_weekly — see comment there.
        state = (session.query(ScheduledJobState)
                 .filter_by(task_name="aps_pub_monthly")
                 .with_for_update()
                 .first())
        end_dt = datetime.now(timezone.utc).replace(tzinfo=None)
        if state and state.last_run_at \
                and (end_dt - state.last_run_at) < timedelta(hours=12):
            logger.info("aps_pub_monthly: ran within last 12 h, skipping duplicate fire")
            session.commit()
            return

        recipients = [
            r.email for r in session.query(ScheduledTaskRecipient)
            .filter_by(task_name="aps_pub_monthly").all()
        ]
        if not recipients:
            logger.warning("aps_pub_monthly: no recipients configured, skipping")
            session.commit()
            return

        start_dt = state.last_run_at if (state and state.last_run_at) \
            else end_dt - timedelta(days=30)

        params = SearchRequest(
            report_type="aps_pub",
            start_date=start_dt.strftime("%Y-%m-%d"),
            end_date=end_dt.strftime("%Y-%m-%d"),
        )
        rows, _ = _run_rcsb_query(params)
        if rows:
            rows = _cross_reference_aps_db(rows, session)
        today_str = end_dt.strftime("%Y-%m-%d")

        if not rows:
            send_mail(
                subject=f"[APS Pub Monthly] No new papers ({today_str})",
                body=(
                    f"No new APS publications were found between "
                    f"{start_dt.strftime('%Y-%m-%d')} and {today_str}."
                ),
                to=recipients,
            )
        else:
            df = pd.DataFrame(rows)
            with tempfile.NamedTemporaryFile(suffix=".xlsx", prefix=f"aps_pub_monthly_{today_str}_", delete=False) as tmp:
                xlsx_path = tmp.name
            _write_excel(df, xlsx_path)
            try:
                send_mail(
                    subject=f"[APS Pub Monthly] {len(rows)} new papers ({today_str})",
                    body=(
                        f"APS publication monthly report: {len(rows)} papers found "
                        f"between {start_dt.strftime('%Y-%m-%d')} and {today_str}.\n"
                        f"See attached spreadsheet."
                    ),
                    to=recipients,
                    attachment_path=xlsx_path,
                )
            finally:
                if os.path.exists(xlsx_path):
                    os.unlink(xlsx_path)

        if state is None:
            state = ScheduledJobState(task_name="aps_pub_monthly")
            session.add(state)
        state.last_run_at = end_dt
        session.commit()
        logger.info(f"aps_pub_monthly completed: {len(rows)} results")

    except Exception as e:
        session.rollback()
        logger.exception(f"aps_pub_monthly job failed: {e}")
        send_admin_alert(
            subject=f"[APS Pub Monthly] Job failed — {type(e).__name__}",
            body=f"Error during aps_pub_monthly scheduled job:\n\n{e}",
        )
    finally:
        session.close()


def _strip_hyperlink(val):
    """Convert =HYPERLINK("url", "text") to just the URL for HTML rendering."""
    if not isinstance(val, str):
        return val
    m = _HYPERLINK_RE.match(val)
    return m.group(1) if m else val

def _run_rcsb_query(params: SearchRequest, progress_cb=None):
    """Run RCSB Search API + GraphQL, return parsed rows as list of dicts."""
    report_type = params.report_type.lower()

    start_date, end_date = get_date_range(
        datetime.strptime(params.start_date, "%Y-%m-%d").date() if params.start_date else None,
        datetime.strptime(params.end_date, "%Y-%m-%d").date() if params.end_date else None,
    )

    query_dict = None
    graphql = None
    parser = None

    if report_type == "aps":
        graphql = Presets.aps_graphql_statement
        parser = APS_Parser(name="APS")
        query_dict = Presets.get_aps_query(start_date, end_date=end_date, APS=True)
    elif report_type == "aps_pub":
        graphql = Presets.publication_graphql_statement
        parser = APS_Pub_Parser(name="APSPub")
        query_dict = Presets.get_aps_query(start_date, end_date=end_date, APS=True)
    elif report_type == "gmca":
        graphql = Presets.aps_graphql_statement
        parser = APS_Parser(name="GMCA")
        query_dict = Presets.get_aps_query(start_date, end_date=end_date, APS=False)
    else:
        # generic — uses Report's default graphql
        parser = Parser(name="report")
        temp_report = Report()
        graphql = temp_report.graph_ql

    # Build PDB ID list
    rcsb = RCSB()
    if query_dict:
        rcsb.search(query_dict)
    if params.search_text:
        rcsb.search_by_text(params.search_text)
    if params.pdb_codes:
        for code in params.pdb_codes:
            rcsb.add(code.strip())

    if not rcsb.pdb_ids:
        return [], parser

    # Batch GraphQL query
    BATCH_SIZE = 200
    total_ids = len(rcsb.pdb_ids)
    all_entries = []
    for i in range(0, total_ids, BATCH_SIZE):
        batch = rcsb.pdb_ids[i:i + BATCH_SIZE]
        if progress_cb:
            progress_cb(f"Fetching entries {i + 1}–{min(i + BATCH_SIZE, total_ids)} of {total_ids}…")
        result = RCSB.query_graphql(graphql % json.dumps(batch))
        if result and "data" in result and "entries" in result["data"]:
            all_entries.extend(result["data"]["entries"])
        if i + BATCH_SIZE < total_ids:
            time.sleep(0.2)

    # Parse entries
    rows = []
    for entry in all_entries:
        try:
            row = parser.parse(entry)
            if hasattr(row, 'release_date') and row.release_date:
                if start_date <= row.release_date <= end_date:
                    rows.append(row._asdict())
            else:
                rows.append(row._asdict())
        except Exception as e:
            logger.warning(f"Failed to parse entry: {e}")
            continue

    return rows, parser


_PREPRINT_JOURNALS = {'biorxiv', 'chemrxiv', 'unpublished', 'not published', 'tba'}


def _postprocess_aps_pub_rows(rows: list) -> list:
    """Group by article (title + article_year), filter preprints/TBD, in_aps_db = any."""
    if not rows:
        return rows

    df = pd.DataFrame(rows).astype(object).fillna("")

    # Filter unpublished / preprint entries
    df = df[~df['title'].isin(['To be published.', 'TBD', ''])]
    if 'journal_abbrev' in df.columns:
        df = df[~df['journal_abbrev'].str.lower().isin(_PREPRINT_JOURNALS)]

    if df.empty:
        return []

    groupby_cols = ['title', 'article_year']

    def _join_unique(s):
        return ",".join(sorted({str(v) for v in s if str(v) not in ('', 'nan', 'None')}))

    agg: dict = {c: _join_unique for c in df.columns if c not in groupby_cols and c != 'in_aps_db'}
    agg['in_aps_db'] = 'any'

    grouped = df.groupby(groupby_cols, as_index=False).agg(agg)
    grouped['in_aps_db'] = grouped['in_aps_db'].astype(bool)
    grouped.sort_values(by=['article_year'], ascending=False, inplace=True, ignore_index=True)

    # in_aps_db first, then rest of columns in original order
    orig_order = ['in_aps_db'] + [c for c in df.columns if c != 'in_aps_db']
    grouped = grouped[[c for c in orig_order if c in grouped.columns]]

    return grouped.to_dict(orient='records')


def _cross_reference_aps_db(rows: list, db_session: Session):
    """Add 'in_aps_db' flag to each row by checking APSPublication table."""
    # Pre-build lookup sets once — O(1) per row instead of one DB query per row.
    doi_set = {r[0] for r in db_session.query(APSPublication.doi).filter(APSPublication.doi.isnot(None)).all()}
    pubmed_set = {r[0] for r in db_session.query(APSPublication.pubmed_id).filter(APSPublication.pubmed_id.isnot(None)).all()}
    title_candidates = [r[0] for r in db_session.query(APSPublication.title).filter(APSPublication.title.isnot(None)).limit(500).all()]

    for row in rows:
        doi = row.get("doi")
        pubmed = row.get("pubmed")
        title = row.get("title", "")

        found = (
            (bool(doi) and doi in doi_set)
            or (bool(pubmed) and str(pubmed) in pubmed_set)
        )

        if not found and title and len(title) > 20:
            for candidate_title in title_candidates:
                if fuzzy_str_compare(title, candidate_title, score_cutoff=95) > 0:
                    found = True
                    break

        row["in_aps_db"] = found

    return rows


def _download_aps_pubdb(db_session: Session):
    """Download APS pub database, solving captcha, and upsert into SQLite."""
    start_url = "https://beam.aps.anl.gov/pls/apsweb/pub_v2_0006.review_start_page"
    query_url = "https://beam.aps.anl.gov/pls/apsweb/pub_v2_0006.query_results"

    http_session = requests.Session()

    # Step 1: GET start page, extract captcha
    try:
        resp = http_session.get(start_url, timeout=30)
        resp.raise_for_status()
    except Exception as e:
        raise HTTPException(status_code=502, detail=f"Failed to reach APS pub database: {e}")

    captcha_session_match = re.search(r'name="p_captcha_session"\s+value="(\d+)"', resp.text)
    if not captcha_session_match:
        raise HTTPException(status_code=502, detail="Could not parse captcha session from APS page")
    captcha_session = captcha_session_match.group(1)

    math_match = re.search(r'What is (\d+)\s*\+\s*(\d+)\s*\?', resp.text)
    if not math_match:
        raise HTTPException(status_code=502, detail="Could not parse captcha math question")
    captcha_answer = int(math_match.group(1)) + int(math_match.group(2))

    # Step 2: POST with captcha answer and export params
    data = {
        "p_captcha_session": captcha_session,
        "p_captcha": str(captcha_answer),
        "i_full_data": "Y",
        "i_non_aps_records": "N",
        "i_order_by_radio": "DEFAULT",
        "i_pend_pub_status_src": "PUBLISHED",
        "i_publication_title": "",
        "i_authorname_last": "",
        "i_authorname_first": "",
        "i_institution": "",
        "i_state": "",
        "i_anl_division": "",
        "i_type_of_publication": "",
        "i_tech_report": "",
        "i_dis_education": "",
        "i_ddl_journal_id1": "",
        "i_ddl_journal_id2": "",
        "i_ddl_journal_id3": "",
        "i_discipline_m": "",
        "i_fund_source": "",
        "i_work_done_cat": "",
        "i_beamline": "",
        "i_gu_beamline": "",
        "i_endstation": "",
        "i_instrument": "",
        "i_entry_month": "",
        "i_entry_year": "",
        "i_entry_from_month": "",
        "i_entry_from_year": "",
        "i_entry_to_month": "",
        "i_entry_to_year": "",
        "i_year_from": "",
        "i_year_to": "",
        "i_doi": "",
        "i_pmid": "",
        "i_pmcid": "",
        "i_pdb": "",
        "i_gup_id": "",
        "i_pup_id": "",
        "i_ipd": "",
        "i_entrydate_from": "",
        "i_entrydate_to": "",
        "i_grant_no": "",
        "i_email_address": "",
        "i_pid": "",
        "i_operation": " Export Excel File ",
    }

    try:
        resp = http_session.post(query_url, data=data, timeout=120)
        resp.raise_for_status()
    except Exception as e:
        raise HTTPException(status_code=502, detail=f"Failed to download APS pub data: {e}")

    # Step 3: Parse CSV response
    content = resp.content.decode(encoding='iso-8859-1')
    try:
        df = pd.read_csv(io.StringIO(content), dtype=str, encoding='iso-8859-1')
    except Exception as e:
        raise HTTPException(status_code=502, detail=f"Failed to parse APS CSV: {e}")

    return _upsert_aps_csv(df, db_session)


def _upsert_aps_csv(df: pd.DataFrame, db_session: Session) -> int:
    """Replace all APSPublication records with rows from a parsed CSV DataFrame."""
    now = datetime.now(timezone.utc).replace(tzinfo=None)
    db_session.query(APSPublication).delete()

    def _clean(val):
        """Convert pandas NaN/NaT to None, everything else to stripped string."""
        if pd.isna(val):
            return None
        s = str(val).strip()
        return s if s else None

    count = 0
    for _, row in df.iterrows():
        pub = APSPublication(
            title=_clean(row.get("Pub Title")),
            doi=_clean(row.get("Doi")),
            pubmed_id=_clean(row.get("Pubmed#")),
            year=_clean(row.get("Year Published")),
            authors=_clean(row.get("Authors")),
            beamline=_clean(row.get("Beamline")),
            journal=_clean(row.get("Journal")),
            volume=_clean(row.get("Volume")),
            pages=_clean(row.get("Page")),
            synced_at=now,
        )
        db_session.add(pub)
        count += 1

    # Update sync status
    sync_status = db_session.query(APSSyncStatus).first()
    if not sync_status:
        sync_status = APSSyncStatus()
        db_session.add(sync_status)
    sync_status.last_sync = now
    sync_status.record_count = count
    sync_status.status = "success"
    sync_status.error_message = None

    db_session.flush()
    logger.info(f"APS pub database synced: {count} records")
    return count


# --- Endpoints ---

def _search_job_worker(job_id: str, params: SearchRequest) -> None:
    """Background thread: run RCSB query, store results in _jobs."""
    session = _get_session()
    try:
        def progress(msg: str):
            _update_job(job_id, progress=msg)

        progress("Querying RCSB Search API…")
        rows, _ = _run_rcsb_query(params, progress_cb=progress)

        if not rows:
            _update_job(job_id, status="done", results=[], count=0, progress="Complete: 0 results")
            return

        if params.report_type.lower() == "aps_pub":
            progress(f"Cross-referencing {len(rows)} entries with APS database…")
            rows = _cross_reference_aps_db(rows, session)
            progress("Grouping by article and filtering preprints…")
            rows = _postprocess_aps_pub_rows(rows)

        progress("Formatting results…")
        for row in rows:
            for key, val in row.items():
                if isinstance(val, date):
                    row[key] = val.isoformat()
                elif key in ("doi", "pubmed"):
                    row[key] = _strip_hyperlink(val)

        _update_job(job_id, status="done", results=rows, count=len(rows),
                    progress=f"Complete: {len(rows)} results")
    except Exception as e:
        logger.exception(f"Search job {job_id} failed: {e}")
        _update_job(job_id, status="error", error=str(e), progress="")
    finally:
        session.close()
        _cleanup_old_jobs()


@router.post("/search")
def search_rcsb(
    params: SearchRequest,
    user: str = Depends(require_staff),
):
    """Start a background RCSB query, return a job_id to poll."""
    job_id = str(uuid.uuid4())
    with _jobs_lock:
        _jobs[job_id] = {"status": "running", "progress": "Starting…", "created_at": time.time()}
    thread = threading.Thread(target=_search_job_worker, args=(job_id, params), daemon=True)
    thread.start()
    return {"job_id": job_id, "status": "running"}


@router.get("/search/status/{job_id}")
def search_status(job_id: str, user: str = Depends(require_staff)):
    """Poll the status of a background search job."""
    with _jobs_lock:
        job = _jobs.get(job_id)
    if not job:
        raise HTTPException(status_code=404, detail="Job not found or expired")
    return job


@router.post("/export")
def export_rcsb(
    params: SearchRequest,
    user: str = Depends(require_staff),
    session: Session = Depends(get_db_session),
):
    """Generate Excel file from RCSB query results."""
    rows, parser = _run_rcsb_query(params)

    if not rows:
        raise HTTPException(status_code=404, detail="No results to export")

    if params.report_type.lower() == "aps_pub":
        rows = _cross_reference_aps_db(rows, session)
        rows = _postprocess_aps_pub_rows(rows)

    df = pd.DataFrame(rows)
    output = io.BytesIO()
    today_str = datetime.now().strftime("%m-%d-%Y")
    report_type = params.report_type.upper()
    filename = f"{report_type}stats-{today_str}.xlsx"
    _write_excel(df, output)
    output.seek(0)

    headers = {"Content-Disposition": f'attachment; filename="{filename}"'}
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers=headers,
    )


@router.post("/sync-aps-db")
def sync_aps_db(
    user: str = Depends(require_staff),
    session: Session = Depends(get_db_session),
):
    """Manually trigger APS pub database sync."""
    try:
        count = _download_aps_pubdb(session)
        return {"status": "ok", "record_count": count}
    except HTTPException:
        raise
    except Exception as e:
        sync_status = session.query(APSSyncStatus).first()
        if not sync_status:
            sync_status = APSSyncStatus()
            session.add(sync_status)
        sync_status.last_sync = datetime.now(timezone.utc).replace(tzinfo=None)
        sync_status.status = "failed"
        sync_status.error_message = str(e)
        session.flush()
        raise HTTPException(status_code=500, detail=f"Sync failed: {e}")


@router.post("/upload-aps-db")
async def upload_aps_db(
    file: UploadFile = File(...),
    user: str = Depends(require_staff),
    session: Session = Depends(get_db_session),
):
    """Upload APS pub database CSV manually (fallback if captcha sync fails)."""
    if not file.filename.endswith(('.csv', '.CSV')):
        raise HTTPException(status_code=400, detail="Only CSV files are accepted")

    content = await file.read()
    try:
        text = content.decode('iso-8859-1')
        df = pd.read_csv(io.StringIO(text), dtype=str, encoding='iso-8859-1')
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Failed to parse CSV: {e}")

    required_cols = {"Pub Title", "Doi"}
    if not required_cols.issubset(set(df.columns)):
        raise HTTPException(
            status_code=400,
            detail=f"CSV missing required columns. Expected at least: {required_cols}. Got: {list(df.columns)[:10]}"
        )

    try:
        count = _upsert_aps_csv(df, session)
        return {"status": "ok", "record_count": count, "source": "upload"}
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Upload failed: {e}")


@router.get("/aps-db-status")
def aps_db_status(
    user: str = Depends(require_staff),
    session: Session = Depends(get_db_session),
):
    """Return last sync time and record count."""
    sync_status = session.query(APSSyncStatus).first()
    if not sync_status:
        return {"last_sync": None, "record_count": 0, "status": "never"}
    return {
        "last_sync": sync_status.last_sync.isoformat() if sync_status.last_sync else None,
        "record_count": sync_status.record_count,
        "status": sync_status.status,
        "error_message": sync_status.error_message,
    }


# --- Scheduled task recipient management ---

@router.get("/scheduled-recipients")
def get_scheduled_recipients(
    user: str = Depends(require_staff),
    session: Session = Depends(get_db_session),
):
    """Return recipient lists for all scheduled tasks."""
    result = {}
    for task_name in _SCHEDULED_TASKS:
        rows = session.query(ScheduledTaskRecipient).filter_by(task_name=task_name).all()
        result[task_name] = [r.email for r in rows]
    return result


@router.put("/scheduled-recipients/{task_name}")
def set_scheduled_recipients(
    task_name: str,
    emails: List[str],
    user: str = Depends(require_staff),
    session: Session = Depends(get_db_session),
):
    """Replace the full recipient list for a scheduled task."""
    if task_name not in _SCHEDULED_TASKS:
        raise HTTPException(status_code=404, detail=f"Unknown task: {task_name}")

    invalid = [e for e in emails if not _EMAIL_RE.match(e)]
    if invalid:
        raise HTTPException(
            status_code=400,
            detail=f"Invalid email address(es): {', '.join(invalid)}",
        )

    session.query(ScheduledTaskRecipient).filter_by(task_name=task_name).delete()
    for email in emails:
        session.add(ScheduledTaskRecipient(task_name=task_name, email=email))
    session.commit()
    return {"task_name": task_name, "recipients": emails}
